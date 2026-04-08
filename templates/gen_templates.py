from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00
import os

out_dir = "/home/claude/risk-dashboard/templates"
os.makedirs(out_dir, exist_ok=True)

# ── Colour palette ──────────────────────────────────────────────
DARK_BLUE   = "1E3A5F"
MID_BLUE    = "2D5F8A"
LIGHT_BLUE  = "BDD7EE"
GOLD        = "C9A84C"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F2F2F2"
MED_GREY    = "D9D9D9"
GREEN_H     = "E2EFDA"
YELLOW_H    = "FFEB9C"
ORANGE_H    = "FFCC99"
RED_H       = "FFC7CE"
GREEN_F     = "375623"
YELLOW_F    = "9C6500"
ORANGE_F    = "974706"
RED_F       = "9C0006"

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_wrap():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def style_header_row(ws, row, cols, bg=DARK_BLUE, fg=WHITE, size=10):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font   = hdr_font(size=size)
        cell.fill   = fill(bg)
        cell.alignment = center()
        cell.border = thin_border()

def style_data_row(ws, row, cols, bg=WHITE):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = body_font()
        cell.fill      = fill(bg)
        cell.alignment = left_wrap()
        cell.border    = thin_border()

# ────────────────────────────────────────────────────────────────
# TEMPLATE 1 – Operational Risk Register
# ────────────────────────────────────────────────────────────────
def build_operational_register():
    wb = Workbook()

    # ── Cover sheet ──────────────────────────────────────────────
    cover = wb.active
    cover.title = "Cover"
    cover.sheet_view.showGridLines = False

    cover.merge_cells("B2:J2")
    t = cover["B2"]
    t.value = "OPERATIONAL RISK REGISTER"
    t.font  = Font(name="Arial", size=22, bold=True, color=WHITE)
    t.fill  = fill(DARK_BLUE)
    t.alignment = center()
    cover.row_dimensions[2].height = 50

    for row_num, (label, val) in enumerate([
        ("Organization / Department:", ""),
        ("Period / Year:",             ""),
        ("Prepared By:",               ""),
        ("Contact Person:",            ""),
        ("Date Prepared:",             ""),
        ("Version:",                   "1.0"),
    ], start=4):
        cover.merge_cells(f"B{row_num}:D{row_num}")
        lbl = cover.cell(row=row_num, column=2, value=label)
        lbl.font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
        lbl.fill = fill(LIGHT_BLUE)
        lbl.alignment = left_wrap()
        lbl.border = thin_border()

        cover.merge_cells(f"E{row_num}:J{row_num}")
        val_cell = cover.cell(row=row_num, column=5, value=val)
        val_cell.font = Font(name="Arial", size=11, color="000000")
        val_cell.fill = fill(LIGHT_GREY)
        val_cell.alignment = left_wrap()
        val_cell.border = thin_border()
        cover.row_dimensions[row_num].height = 22

    for col_l, w in [("A",3),("B",28),("C",4),("D",4),("E",20),("F",10),("G",10),("H",10),("I",10),("J",10)]:
        set_col_width(cover, col_l, w)

    # ── Rating scale legend ──────────────────────────────────────
    legend = wb.create_sheet("Rating Scale")
    legend.sheet_view.showGridLines = False

    legend.merge_cells("B2:H2")
    title = legend["B2"]
    title.value = "RISK RATING SCALE"
    title.font  = hdr_font(size=14)
    title.fill  = fill(DARK_BLUE)
    title.alignment = center()
    legend.row_dimensions[2].height = 36

    # Likelihood table
    legend.merge_cells("B4:H4")
    lt = legend["B4"]
    lt.value = "LIKELIHOOD SCALE"
    lt.font  = hdr_font(size=11)
    lt.fill  = fill(MID_BLUE)
    lt.alignment = center()

    l_headers = ["Score","Level","Description"]
    l_data    = [
        (1,"Rare",         "May occur only in exceptional circumstances (< once in 10 years)"),
        (2,"Unlikely",     "Could occur at some time (once in 5–10 years)"),
        (3,"Moderate",     "Might occur at some time (once in 1–5 years)"),
        (4,"Likely",       "Will probably occur in most circumstances (once a year)"),
        (5,"Common",       "Expected to occur in most circumstances (more than once a year)"),
    ]
    for ci, h in enumerate(l_headers, start=2):
        c = legend.cell(row=5, column=ci, value=h)
        c.font = hdr_font(size=10); c.fill = fill(LIGHT_BLUE)
        c.alignment = center(); c.border = thin_border()
        c.font = Font(name="Arial", size=10, bold=True, color=DARK_BLUE)

    for ri, row_data in enumerate(l_data, start=6):
        for ci, v in enumerate(row_data, start=2):
            c = legend.cell(row=ri, column=ci, value=v)
            c.font   = body_font()
            c.fill   = fill(LIGHT_GREY if ri % 2 == 0 else WHITE)
            c.alignment = left_wrap() if ci == 4 else center()
            c.border = thin_border()
        legend.row_dimensions[ri].height = 20

    # Impact table
    legend.merge_cells("B12:H12")
    it = legend["B12"]
    it.value = "IMPACT SCALE"
    it.font  = hdr_font(size=11)
    it.fill  = fill(MID_BLUE)
    it.alignment = center()

    i_headers = ["Score","Level","Description"]
    i_data    = [
        (1,"Insignificant","Negligible financial or operational impact"),
        (2,"Minor",        "Small financial loss; minor disruption < 1 day"),
        (3,"Moderate",     "Moderate loss; disruption 1–7 days; manageable"),
        (4,"Major",        "Significant loss; disruption > 1 week; reputational damage"),
        (5,"Critical",     "Catastrophic loss; long-term disruption; regulatory action"),
    ]
    for ci, h in enumerate(i_headers, start=2):
        c = legend.cell(row=13, column=ci, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=DARK_BLUE)
        c.fill = fill(LIGHT_BLUE); c.alignment = center(); c.border = thin_border()

    for ri, row_data in enumerate(i_data, start=14):
        for ci, v in enumerate(row_data, start=2):
            c = legend.cell(row=ri, column=ci, value=v)
            c.font   = body_font()
            c.fill   = fill(LIGHT_GREY if ri % 2 == 0 else WHITE)
            c.alignment = left_wrap() if ci == 4 else center()
            c.border = thin_border()
        legend.row_dimensions[ri].height = 20

    # Risk rating matrix key
    legend.merge_cells("B20:H20")
    mk = legend["B20"]
    mk.value = "RISK RATING = LIKELIHOOD × IMPACT"
    mk.font  = hdr_font(size=11); mk.fill = fill(MID_BLUE); mk.alignment = center()

    rating_key = [
        ("1 – 4",  "Low",      GREEN_H,  GREEN_F),
        ("5 – 9",  "Tolerable",YELLOW_H, YELLOW_F),
        ("10 – 16","High",     ORANGE_H, ORANGE_F),
        ("17 – 25","Critical", RED_H,    RED_F),
    ]
    for ci, h in enumerate(["Score Range","Rating","Action Required"], start=2):
        c = legend.cell(row=21, column=ci, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=DARK_BLUE)
        c.fill = fill(LIGHT_BLUE); c.alignment = center(); c.border = thin_border()

    actions = [
        "Monitor; review annually",
        "Review periodically; assign owner",
        "Immediate management attention; mitigation required",
        "Escalate to senior management; urgent action required",
    ]
    for ri, (score, rating, bg, fg) in enumerate(rating_key, start=22):
        for ci, val in enumerate([score, rating, actions[ri-22]], start=2):
            c = legend.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10, bold=True, color=fg)
            c.fill = fill(bg); c.alignment = center(); c.border = thin_border()
        legend.row_dimensions[ri].height = 20

    for col_l, w in [("A",3),("B",10),("C",16),("D",50),("E",10),("F",10),("G",10),("H",10)]:
        set_col_width(legend, col_l, w)

    # ── Operational Risk Register sheet ─────────────────────────
    ws = wb.create_sheet("Operational Risk Register")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"

    # Title block
    ws.merge_cells("A1:U1")
    ws["A1"].value = "OPERATIONAL RISK REGISTER"
    ws["A1"].font  = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill  = fill(DARK_BLUE)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 40

    info_labels = ["Department / Unit:", "Period:", "Prepared By:", "Contact:"]
    for i, lbl in enumerate(info_labels):
        col = 1 + i * 5
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        ws.merge_cells(start_row=2, start_column=col+2, end_row=2, end_column=col+3)
        lc = ws.cell(row=2, column=col, value=lbl)
        lc.font = Font(name="Arial", size=9, bold=True, color=DARK_BLUE)
        lc.fill = fill(LIGHT_BLUE); lc.alignment = left_wrap(); lc.border = thin_border()
        vc = ws.cell(row=2, column=col+2, value="")
        vc.fill = fill(LIGHT_GREY); vc.border = thin_border(); vc.alignment = left_wrap()
    ws.row_dimensions[2].height = 18

    # Section group headers (row 8)
    group_headers = [
        (1,  3,  "RISK IDENTIFICATION"),
        (4,  5,  "INHERENT RISK"),
        (6,  6,  "SCORE"),
        (7,  7,  "RATING"),
        (8,  10, "CONTROLS"),
        (11, 12, "RESIDUAL RISK"),
        (13, 13, "RES. SCORE"),
        (14, 14, "RES. RATING"),
        (15, 17, "TREATMENT PLAN"),
        (18, 19, "ACCOUNTABILITY"),
        (20, 21, "MONITORING"),
    ]
    for (sc, ec, label) in group_headers:
        if sc == ec:
            c = ws.cell(row=8, column=sc, value=label)
        else:
            ws.merge_cells(start_row=8, start_column=sc, end_row=8, end_column=ec)
            c = ws.cell(row=8, column=sc, value=label)
        c.font = hdr_font(size=9); c.fill = fill(MID_BLUE); c.alignment = center(); c.border = thin_border()
    ws.row_dimensions[8].height = 22

    # Column headers (row 9)
    col_headers = [
        "Risk ID", "Key Business Process / Objective", "Risk Description",
        "Likelihood\n(1–5)", "Impact\n(1–5)", "Inherent\nRisk Score", "Inherent\nRisk Rating",
        "Existing Controls", "Control Effectiveness", "Control Rating",
        "Residual\nLikelihood", "Residual\nImpact", "Residual\nRisk Score", "Residual\nRisk Rating",
        "Treatment\nOption", "Treatment Actions\n(SMART)", "Timeframe",
        "Risk Owner", "Review Date",
        "Status", "Remarks",
    ]
    for ci, h in enumerate(col_headers, start=1):
        c = ws.cell(row=9, column=ci, value=h)
        c.font = hdr_font(size=9); c.fill = fill(DARK_BLUE)
        c.alignment = center(); c.border = thin_border()
    ws.row_dimensions[9].height = 40

    # Sample data rows
    samples = [
        ("OPR-001","Procurement","Risk of supplier failure to deliver goods on time",3,4,None,"",
         "Dual supplier policy; quarterly supplier reviews","Partially Effective","Partial",2,3,None,"",
         "Treat","Identify and onboard backup suppliers by Q2","Short-term","Procurement Manager","2026-06-30","Open",""),
        ("OPR-002","IT Infrastructure","Risk of cybersecurity breach compromising data",4,5,None,"",
         "Firewall; antivirus; annual penetration testing","Effective","Strong",2,4,None,"",
         "Treat","Implement zero-trust architecture; staff training","Medium-term","IT Manager","2026-09-30","In Progress",""),
        ("OPR-003","Financial Reporting","Risk of material misstatement in financial statements",2,5,None,"",
         "Internal audit; external audit; segregation of duties","Effective","Strong",1,4,None,"",
         "Tolerate","Annual external audit; strengthen internal controls","Ongoing","CFO","2026-12-31","Monitoring",""),
        ("OPR-004","Human Resources","Risk of key staff turnover disrupting operations",3,3,None,"",
         "Competitive remuneration; succession planning","Partially Effective","Partial",2,2,None,"",
         "Treat","Develop succession plans for all critical roles","Long-term","HR Director","2026-06-30","Open",""),
        ("OPR-005","Compliance","Risk of regulatory non-compliance resulting in penalties",2,4,None,"",
         "Compliance calendar; legal reviews","Effective","Strong",1,3,None,"",
         "Treat","Quarterly compliance audits; update policy register","Ongoing","Legal Officer","2026-03-31","In Progress",""),
    ]

    RATING_FILLS = {
        "Low":      (GREEN_H,  GREEN_F),
        "Tolerable":(YELLOW_H, YELLOW_F),
        "High":     (ORANGE_H, ORANGE_F),
        "Critical": (RED_H,    RED_F),
    }

    def score_rating(score):
        if score <= 4:  return "Low"
        if score <= 9:  return "Tolerable"
        if score <= 16: return "High"
        return "Critical"

    for ri, row_data in enumerate(samples, start=10):
        bg = LIGHT_GREY if ri % 2 == 0 else WHITE
        (rid,kbp,rdesc,il,ii,_,_,ctrl,ctrl_eff,ctrl_rat,rl,ri2,_,_,topt,tact,tf,owner,rdate,status,rem) = row_data
        inh_score = il * ii
        res_score = rl * ri2
        inh_rat = score_rating(inh_score)
        res_rat = score_rating(res_score)

        row_vals = [rid,kbp,rdesc,il,ii,inh_score,inh_rat,ctrl,ctrl_eff,ctrl_rat,rl,ri2,res_score,res_rat,topt,tact,tf,owner,rdate,status,rem]
        for ci, val in enumerate(row_vals, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = body_font(size=9); c.alignment = left_wrap(); c.border = thin_border()

            if ci in (4,5,11,12):  # numeric ratings – center
                c.alignment = center()
            if ci == 6:  # inherent score – color
                rfill, rfont = RATING_FILLS[inh_rat]
                c.fill = fill(rfill); c.font = Font(name="Arial", size=9, bold=True, color=rfont)
                c.alignment = center()
            elif ci == 7:  # inherent rating
                rfill, rfont = RATING_FILLS[inh_rat]
                c.fill = fill(rfill); c.font = Font(name="Arial", size=9, bold=True, color=rfont)
                c.alignment = center()
            elif ci == 13:  # residual score
                rfill, rfont = RATING_FILLS[res_rat]
                c.fill = fill(rfill); c.font = Font(name="Arial", size=9, bold=True, color=rfont)
                c.alignment = center()
            elif ci == 14:  # residual rating
                rfill, rfont = RATING_FILLS[res_rat]
                c.fill = fill(rfill); c.font = Font(name="Arial", size=9, bold=True, color=rfont)
                c.alignment = center()
            else:
                c.fill = fill(bg)
        ws.row_dimensions[ri].height = 45

    # Add 10 blank template rows
    for ri in range(15, 26):
        bg = LIGHT_GREY if ri % 2 == 0 else WHITE
        for ci in range(1, 22):
            c = ws.cell(row=ri, column=ci, value="")
            c.font = body_font(size=9); c.fill = fill(bg)
            c.alignment = left_wrap(); c.border = thin_border()
        ws.row_dimensions[ri].height = 35

    # Column widths
    col_widths = [10,28,32,12,12,12,14,28,22,14,12,12,12,14,14,32,18,18,14,14,20]
    for ci, w in enumerate(col_widths, start=1):
        set_col_width(ws, get_column_letter(ci), w)

    path = os.path.join(out_dir, "operational_risk_register_template.xlsx")
    wb.save(path)
    print(f"✓ Saved: {path}")
    return path

# ────────────────────────────────────────────────────────────────
# TEMPLATE 2 – Strategic Risk Register (simpler, different style)
# ────────────────────────────────────────────────────────────────
def build_strategic_register():
    wb = Workbook()
    ws = wb.active
    ws.title = "Strategic Risk Register"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"

    # Title
    ws.merge_cells("A1:S1")
    ws["A1"].value = "STRATEGIC RISK REGISTER"
    ws["A1"].font  = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill  = fill("2C3E50")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 40

    # Subtitle info row
    for col, lbl, val in [(1,"Organization:",""), (5,"Period:",""), (9,"Risk Appetite:","Medium"), (13,"Version:","1.0")]:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        ws.merge_cells(start_row=2, start_column=col+2, end_row=2, end_column=col+3)
        lc = ws.cell(row=2, column=col, value=lbl)
        lc.font = Font(name="Arial", size=9, bold=True, color="2C3E50")
        lc.fill = fill("AED6F1"); lc.alignment = left_wrap(); lc.border = thin_border()
        vc = ws.cell(row=2, column=col+2, value=val)
        vc.fill = fill(LIGHT_GREY); vc.border = thin_border(); vc.alignment = left_wrap()
    ws.row_dimensions[2].height = 18

    # Column headers
    headers = [
        "Risk\nID", "Strategic\nObjective", "Risk\nDescription", "Risk\nCategory",
        "Causes", "Consequence", "Inherent\nLikelihood", "Inherent\nImpact",
        "Inherent\nScore", "Inherent\nRating", "Controls\nin Place", "Control\nRating",
        "Residual\nLikelihood", "Residual\nImpact", "Residual\nScore", "Residual\nRating",
        "Treatment\nStrategy", "Risk Owner", "Target Date",
    ]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=9, column=ci, value=h)
        c.font = hdr_font(size=9); c.fill = fill("2C3E50")
        c.alignment = center(); c.border = thin_border()
    ws.row_dimensions[9].height = 40

    strat_samples = [
        ("STR-001","Market Growth","Risk of losing market share to competitors",3,4,"Economic","Competitive pressure","Revenue decline"),
        ("STR-002","Digital Transform.","Risk of failure to implement digital strategy on time",4,4,"Technological","Resource constraints","Delayed benefits"),
        ("STR-003","Stakeholder Rel.","Risk of damage to key stakeholder relationships",2,5,"Reputational","Communication gaps","Loss of funding"),
        ("STR-004","Regulatory Env.","Risk of adverse regulatory changes",3,4,"Legal","Policy uncertainty","Compliance costs"),
        ("STR-005","Talent Strategy","Risk of failing to attract and retain top talent",3,3,"Strategic","Labour market","Capacity gaps"),
    ]

    def score_rating(s):
        if s<=4: return "Low"
        if s<=9: return "Tolerable"
        if s<=16: return "High"
        return "Critical"

    FILLS = {"Low":(GREEN_H,GREEN_F),"Tolerable":(YELLOW_H,YELLOW_F),"High":(ORANGE_H,ORANGE_F),"Critical":(RED_H,RED_F)}

    for ri, (rid,obj,desc,il,ii,cat,cause,cons) in enumerate(strat_samples, start=10):
        inh = il*ii; res_l=max(1,il-1); res_i=max(1,ii-1); res=res_l*res_i
        irat = score_rating(inh); rrat = score_rating(res)
        bg = LIGHT_GREY if ri % 2 == 0 else WHITE

        row_vals = [rid,obj,desc,cat,cause,cons,il,ii,inh,irat,"Existing controls","Partial",res_l,res_i,res,rrat,"Treat","Risk Owner","2026-12-31"]
        for ci, val in enumerate(row_vals, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = body_font(size=9); c.fill = fill(bg)
            c.alignment = center() if ci in (7,8,9,10,12,13,14,15,16) else left_wrap()
            c.border = thin_border()
            if ci in (9,10):
                rfill, rfont = FILLS[irat]
                c.fill = fill(rfill); c.font = Font(name="Arial", size=9, bold=True, color=rfont)
            if ci in (15,16):
                rfill, rfont = FILLS[rrat]
                c.fill = fill(rfill); c.font = Font(name="Arial", size=9, bold=True, color=rfont)
        ws.row_dimensions[ri].height = 40

    for ri in range(15, 26):
        bg = LIGHT_GREY if ri % 2 == 0 else WHITE
        for ci in range(1, 20):
            c = ws.cell(row=ri, column=ci, value="")
            c.font = body_font(size=9); c.fill = fill(bg)
            c.alignment = left_wrap(); c.border = thin_border()
        ws.row_dimensions[ri].height = 35

    col_widths = [10,22,30,18,22,22,12,12,12,14,24,14,12,12,12,14,16,18,14]
    for ci, w in enumerate(col_widths, start=1):
        set_col_width(ws, get_column_letter(ci), w)

    path = os.path.join(out_dir, "strategic_risk_register_template.xlsx")
    wb.save(path)
    print(f"✓ Saved: {path}")
    return path

# ────────────────────────────────────────────────────────────────
# TEMPLATE 3 – Simple CSV Template
# ────────────────────────────────────────────────────────────────
def build_csv_template():
    import csv
    path = os.path.join(out_dir, "risk_import_template.csv")
    headers = [
        "Risk ID","Key Business Process","Risk Description","Category",
        "Causes","Consequence",
        "Likelihood","Impact",
        "Controls","Control Rating",
        "Residual Likelihood","Residual Impact",
        "Treatment Option","Treatment Actions","Timeframe",
        "Risk Owner","Status",
    ]
    rows = [
        ["OPR-001","Procurement","Supplier delivery failure","Operational",
         "Single supplier dependency","Stock-outs disrupting operations",
         "3","4","Dual supplier policy","Partial",
         "2","3","Treat","Onboard backup supplier Q2 2026","Short-term","Procurement Manager","Open"],
        ["FIN-001","Financial Reporting","Material misstatement risk","Financial Management",
         "Weak internal controls","Regulatory penalties",
         "2","5","External audit; internal review","Effective",
         "1","4","Treat","Strengthen reconciliation controls","Ongoing","CFO","In Progress"],
        ["TEC-001","IT Infrastructure","Cybersecurity breach","Technological",
         "Outdated systems; phishing","Data loss; reputational damage",
         "4","5","Firewall; antivirus; pen testing","Partial",
         "2","4","Treat","Zero-trust architecture rollout","Medium-term","IT Manager","In Progress"],
    ]
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)
    print(f"✓ Saved: {path}")
    return path

p1 = build_operational_register()
p2 = build_strategic_register()
p3 = build_csv_template()
print("All templates generated.")
